"""
Statbate Scraper — Flask Backend
"""

import datetime
import io
import json
import os
import random
import re
import tempfile
import threading
import time
import uuid
from pathlib import Path

from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

TOKEN_COST = 0.05

PLATFORM_NAMES = {
    "1": "Chaturbate",
    "2": "BongaCams",
    "3": "Stripchat",
    "4": "CamSoda",
    "6": "MFC",
}

# ── Job persistence ────────────────────────────────────────────────────────────

JOBS_DIR = Path(os.environ.get("JOBS_DIR", "/tmp/statbate_jobs"))
JOBS_DIR.mkdir(parents=True, exist_ok=True)

def _cleanup_orphaned_jobs():
    """On startup, mark any running/queued jobs as stopped — they were killed by a redeploy."""
    for p in JOBS_DIR.glob("*.json"):
        try:
            with open(p) as f:
                job = json.load(f)
            if job.get("status") in ("running", "queued"):
                job["status"] = "stopped"
                job.setdefault("log", []).append("  ⚠ marked stopped (orphaned at startup)")
                with open(p, "w") as f:
                    json.dump(job, f)
        except Exception:
            pass

_cleanup_orphaned_jobs()

def _job_path(job_id):
    return JOBS_DIR / f"{job_id}.json"

def _results_dir(job_id):
    d = JOBS_DIR / f"{job_id}_results"
    d.mkdir(exist_ok=True)
    return d

def load_job(job_id):
    p = _job_path(job_id)
    if p.exists():
        with open(p) as f:
            return json.load(f)
    return None

def save_job(job):
    # If a results folder exists for this job, strip results from JSON (stored per-file)
    # Otherwise keep results in JSON (legacy jobs or jobs not yet started)
    results_dir = JOBS_DIR / f"{job['id']}_results"
    if results_dir.exists():
        meta = {k: v for k, v in job.items() if k != "results"}
    else:
        meta = job
    with open(_job_path(job["id"]), "w") as f:
        json.dump(meta, f)

def append_result(job_id, username, data):
    """Write a single model result to its own file."""
    safe = re.sub(r'[^\w\-]', '_', username)
    p = _results_dir(job_id) / f"{safe}.json"
    with open(p, "w") as f:
        json.dump({"username": username, "data": data}, f)

def load_results(job_id):
    """Load all model results from disk. Falls back to job JSON for legacy jobs."""
    results_dir = JOBS_DIR / f"{job_id}_results"
    if results_dir.exists():
        results = []
        for p in sorted(results_dir.glob("*.json")):
            try:
                with open(p) as f:
                    results.append(json.load(f))
            except Exception:
                pass
        return results
    # Legacy fallback: results were stored inside the job JSON
    job = load_job(job_id)
    if job and job.get("results"):
        return job["results"]
    return []

def completed_usernames(job_id):
    """Return set of usernames already scraped for a job."""
    results_dir = JOBS_DIR / f"{job_id}_results"
    if not results_dir.exists():
        return set()
    done = set()
    for p in results_dir.glob("*.json"):
        try:
            with open(p) as f:
                done.add(json.load(f)["username"])
        except Exception:
            pass
    return done

def list_jobs_from_disk():
    jobs_list = []
    for p in sorted(JOBS_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True)[:20]:
        try:
            with open(p) as f:
                jobs_list.append(json.load(f))
        except Exception:
            pass
    return jobs_list


# ── Styling ────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1A3A5C")
ROW_ODD_FILL  = PatternFill("solid", fgColor="EFF4FB")
ROW_EVEN_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(size=10)
THIN = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)
OK_FILL   = PatternFill("solid", fgColor="D6F5E3")
MISS_FILL = PatternFill("solid", fgColor="FDE8E8")
ERR_FILL  = PatternFill("solid", fgColor="FFF3CD")


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def style_row(ws, row, ncols, odd):
    fill = ROW_ODD_FILL if odd else ROW_EVEN_FILL
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN


def auto_width(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[letter].width = min(max(width + 4, 12), 55)


# ── Scraping ───────────────────────────────────────────────────────────────────

def extract_js_var(html, var_name):
    pattern = rf'var\s+{re.escape(var_name)}\s*=\s*(\[.*?\]|\{{.*?\}}|".*?"|\'.*?\'|-?\d+\.?\d*);'
    match = re.search(pattern, html, re.DOTALL)
    if not match:
        return None
    try:
        return json.loads(match.group(1))
    except Exception:
        return None


def extract_profile_info(page):
    info = {}
    try:
        rows = page.locator(".bg-white table tbody tr").all()
        for row in rows:
            cells = row.locator("td").all()
            if len(cells) >= 2:
                label = cells[0].inner_text().strip().lower()
                value = cells[1].inner_text().strip()
                value = re.sub(r'\s*\?.*', '', value).strip()
                if "type" in label:
                    info["type"] = value
                elif "last online" in label:
                    info["last_online"] = value
                elif "last month" in label:
                    info["last_month_usd"] = value
                elif "all time" in label:
                    info["all_time_usd"] = value
    except Exception as e:
        info["profile_error"] = str(e)
    return info


def scrape_user(page, platform, username):
    url = f"https://statbate.com/search/{platform}/{username}"
    result = {"status": "ok", "url": url, "profile": {}, "charts": {}, "tables": {}}

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(2000)
        html = page.content()

        body_text = page.locator("body").inner_text()
        if any(x in body_text.lower() for x in ["not found", "no data", "404", "doesn't exist"]):
            result["status"] = "not_found"
            return result

        result["profile"] = extract_profile_info(page)

        result["charts"] = {
            "earnings_cumulative": extract_js_var(html, "chartIncomeLine") or [],
            "daily_income":        extract_js_var(html, "chartIncomeBar")  or [],
            "tippers_cumulative":  extract_js_var(html, "chartDonsLine")   or [],
            "tips_count":          extract_js_var(html, "chartTipsLine")   or [],
        }

        api_url = f"https://statbate.com/api.php?action=tables&name={username}&db={platform}"
        api_response = page.evaluate(f"""
            async () => {{
                try {{
                    const r = await fetch('{api_url}');
                    return await r.json();
                }} catch(e) {{ return null; }}
            }}
        """)
        result["tables"] = api_response or {}

    except Exception as e:
        result["status"] = f"error: {e}"

    return result


def scrape_sc_socials(page, username):
    """Visit Stripchat profile page and extract social links from bio."""
    socials = {"twitter": "", "instagram": "", "onlyfans": "", "other": [], "bio": ""}
    try:
        url = f"https://stripchat.com/{username}"
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
        page.wait_for_timeout(2000)

        # Try to get bio text
        try:
            bio_el = page.locator("[class*='about'], [class*='bio'], [class*='description']").first
            if bio_el.count():
                socials["bio"] = bio_el.inner_text().strip()[:500]
        except Exception:
            pass

        # Extract all external links from the page
        try:
            links = page.evaluate("""
                () => Array.from(document.querySelectorAll('a[href]'))
                    .map(a => a.href)
                    .filter(h => h.startsWith('http') && !h.includes('stripchat.com'))
            """)
            for link in links:
                link = link.strip().rstrip('/')
                if not link:
                    continue
                if 'twitter.com' in link or 'x.com' in link:
                    if not socials["twitter"]:
                        socials["twitter"] = link
                elif 'instagram.com' in link:
                    if not socials["instagram"]:
                        socials["instagram"] = link
                elif 'onlyfans.com' in link:
                    if not socials["onlyfans"]:
                        socials["onlyfans"] = link
                elif link not in socials["other"]:
                    socials["other"].append(link)
        except Exception:
            pass

        # Also scan bio text for handles/links not wrapped in <a> tags
        bio = socials["bio"]
        if bio:
            if not socials["twitter"]:
                m = re.search(r'(?:twitter\.com/|x\.com/|@)([A-Za-z0-9_]{1,50})', bio)
                if m:
                    socials["twitter"] = f"https://twitter.com/{m.group(1)}"
            if not socials["instagram"]:
                m = re.search(r'(?:instagram\.com/|ig:\s*@?)([A-Za-z0-9_.]{1,50})', bio, re.IGNORECASE)
                if m:
                    socials["instagram"] = f"https://instagram.com/{m.group(1)}"
            if not socials["onlyfans"]:
                m = re.search(r'onlyfans\.com/([A-Za-z0-9_.]{1,50})', bio, re.IGNORECASE)
                if m:
                    socials["onlyfans"] = f"https://onlyfans.com/{m.group(1)}"

        # Cap other links to 3
        socials["other"] = socials["other"][:3]

    except Exception as e:
        socials["error"] = str(e)

    return socials
    rows = []
    for point in data:
        if len(point) >= 2:
            date_str = datetime.datetime.utcfromtimestamp(point[0] / 1000).strftime("%Y-%m-%d")
            rows.append({"date": date_str, col2_name: point[1]})
    return rows


# ── Job runner ─────────────────────────────────────────────────────────────────

# In-memory stop signals: job_id -> threading.Event
_stop_flags: dict = {}

def run_scrape_job(job_id, platform, usernames, delay):
    job = load_job(job_id)
    job["status"] = "running"
    job["total"] = len(usernames)
    job["done"] = 0
    job["log"] = []
    # preserve usernames list for resume
    job["all_usernames"] = usernames
    save_job(job)

    stop_event = _stop_flags.setdefault(job_id, threading.Event())

    RECYCLE_EVERY = 50

    def make_page(pw):
        browser = pw.chromium.launch(headless=True, args=["--disable-dev-shm-usage"])
        ctx = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
        )
        return browser, ctx.new_page()

    try:
        with sync_playwright() as pw:
            browser, page = make_page(pw)

            for idx, username in enumerate(usernames, 1):
                # Check stop signal
                if stop_event.is_set():
                    job["log"].append("  ⏹ stopped by user.")
                    try:
                        browser.close()
                    except Exception:
                        pass
                    job["status"] = "stopped"
                    save_job(job)
                    return

                # Recycle browser periodically to free memory
                if idx > 1 and (idx - 1) % RECYCLE_EVERY == 0:
                    try:
                        browser.close()
                    except Exception:
                        pass
                    job["log"].append(f"  ↻ recycling browser at model {idx}...")
                    save_job(job)
                    browser, page = make_page(pw)

                job["log"].append(f"[{idx}/{len(usernames)}] Scraping {username}...")
                data = scrape_user(page, platform, username)

                # For Stripchat, also scrape social links from the profile page
                if platform == "3":
                    socials = scrape_sc_socials(page, username)
                    data["socials"] = socials
                    has = [k for k in ("twitter","instagram","onlyfans") if socials.get(k)] + socials.get("other",[])
                    job["log"][-1] += f" {data.get('status','?')} | socials: {len(has)} found"
                else:
                    data["socials"] = {}
                    job["log"][-1] += f" {data.get('status', '?')}"

                append_result(job_id, username, data)
                job["done"] = idx
                save_job(job)

                if idx < len(usernames):
                    time.sleep(random.uniform(delay * 0.7, delay * 1.4))

            browser.close()

        job["status"] = "done"
        save_job(job)

    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        save_job(job)
    finally:
        _stop_flags.pop(job_id, None)


# ── Excel / CSV builders ───────────────────────────────────────────────────────

def _build_excel_workbook(all_data, platform, sections=None):
    ALL = {"profile","last30d","recent_tips","top_monthly","top_alltime","biggest_tips"}
    inc = sections if sections else ALL
    wb = Workbook()
    platform_name = PLATFORM_NAMES.get(str(platform), f"Platform {platform}")

    SOCIAL_HEADERS = ["Username", "Platform", "Status", "Type", "Last Online",
                      "Last Month $", "All Time $", "Twitter/X", "Instagram", "OnlyFans", "Other Links", "Bio"]
    BASE_HEADERS   = ["Username", "Platform", "Status", "Type", "Last Online", "Last Month $", "All Time $"]

    NO_SOCIAL_HEADERS = ["Username", "Platform", "Status", "Type", "Last Online",
                         "Last Month $", "All Time $", "Twitter/X (try)", "Instagram (try)", "OnlyFans (try)"]
    BASE_HEADERS      = ["Username", "Platform", "Status", "Type", "Last Online",
                         "Last Month $", "All Time $", "Twitter/X (try)", "Instagram (try)", "OnlyFans (try)"]

    def model_row_social(entry):
        p = entry["data"].get("profile", {}) if "profile" in inc else {}
        s = entry["data"].get("socials", {})
        return [
            entry["username"], platform_name,
            entry["data"].get("status", ""),
            p.get("type", ""), p.get("last_online", ""),
            p.get("last_month_usd", ""), p.get("all_time_usd", ""),
            s.get("twitter", ""), s.get("instagram", ""), s.get("onlyfans", ""),
            " | ".join(s.get("other", [])), s.get("bio", ""),
        ]

    def model_row_base(entry):
        p = entry["data"].get("profile", {}) if "profile" in inc else {}
        u = entry["username"]
        return [
            u, platform_name,
            entry["data"].get("status", ""),
            p.get("type", ""), p.get("last_online", ""),
            p.get("last_month_usd", ""), p.get("all_time_usd", ""),
            f"https://twitter.com/{u}",
            f"https://instagram.com/{u}",
            f"https://onlyfans.com/{u}",
        ]

    def has_socials(entry):
        s = entry["data"].get("socials", {})
        return bool(s.get("twitter") or s.get("instagram") or s.get("onlyfans") or s.get("other"))

    def make_summary_sheet(title, headers, rows_data, status_col=2):
        ws = wb.create_sheet(title)
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        style_header(ws, 1, len(headers))
        for i, row in enumerate(rows_data, 1):
            r = i + 1
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
            style_row(ws, r, len(headers), i % 2 == 1)
            status = row[status_col]
            sc = ws.cell(row=r, column=status_col + 1)
            if status == "ok": sc.fill = OK_FILL
            elif status == "not_found": sc.fill = MISS_FILL
            elif "error" in str(status): sc.fill = ERR_FILL
        auto_width(ws)
        ws.freeze_panes = "A2"

    # Split into has/no socials only for SC (platform 3)
    if platform == "3":
        with_socials = [e for e in all_data if has_socials(e)]
        without_socials = [e for e in all_data if not has_socials(e)]
        ws_first = wb.active
        wb.remove(ws_first)
        make_summary_sheet("Has Socials", SOCIAL_HEADERS,
                           [model_row_social(e) for e in with_socials])
        make_summary_sheet("No Socials", NO_SOCIAL_HEADERS,
                           [model_row_base(e) for e in without_socials])
    else:
        # Non-SC platforms — single Summary sheet
        ws_active = wb.active
        wb.remove(ws_active)
        make_summary_sheet("Summary", BASE_HEADERS,
                           [model_row_base(e) for e in all_data])

    def make_sheet(title, headers_list, row_fn):
        s = wb.create_sheet(title)
        for c, h in enumerate(headers_list, 1):
            s.cell(row=1, column=c, value=h)
        style_header(s, 1, len(headers_list))
        ri = 2; par = 1
        for entry in all_data:
            rows = row_fn(entry)
            if not rows:
                s.cell(row=ri, column=1, value=entry["username"])
                style_row(s, ri, len(headers_list), par % 2 == 1)
                ri += 1; par += 1; continue
            for row in rows:
                for c, val in enumerate(row, 1):
                    s.cell(row=ri, column=c, value=val)
                style_row(s, ri, len(headers_list), par % 2 == 1)
                ri += 1; par += 1
        auto_width(s)
        s.freeze_panes = "B2"

    if "last30d" in inc:
        make_sheet("Last 30d",
            ["Username", "Date", "Dons", "Tips", "Avg USD", "Total USD"],
            lambda e: [[e["username"], r.get("date",""), r.get("dons",""), r.get("tips",""),
                        f"${r.get('avg_usd','')}", f"${r.get('total_usd','')}"]
                       for r in (e["data"].get("tables",{}).get("searchIncome") or [])])

    if "recent_tips" in inc:
        make_sheet("Recent tips",
            ["Username", "Date", "Donator", "Tokens", "USD"],
            lambda e: [[e["username"], r.get("date",""), r.get("don_name",""),
                        r.get("tokens",""), f"${r.get('usd','')}"]
                       for r in _get_tip_rows(e["data"].get("tables",{}).get("incomeDetails"))])

    if "top_monthly" in inc:
        make_sheet("Top monthly",
            ["Username", "Rank", "Donator", "Tokens", "USD"],
            lambda e: [[e["username"], i+1, r.get("name",""),
                        round(r.get("total_usd",0)/TOKEN_COST) if r.get("total_usd") else "",
                        f"${r.get('total_usd','')}"]
                       for i, r in enumerate(e["data"].get("tables",{}).get("donsMonth") or [])])

    if "top_alltime" in inc:
        make_sheet("Top all-time",
            ["Username", "Rank", "Donator", "Tokens", "USD"],
            lambda e: [[e["username"], i+1, r.get("name",""),
                        round(r.get("total_usd",0)/TOKEN_COST) if r.get("total_usd") else "",
                        f"${r.get('total_usd','')}"]
                       for i, r in enumerate(e["data"].get("tables",{}).get("donsAll") or [])])

    if "biggest_tips" in inc:
        make_sheet("Biggest tips",
            ["Username", "Date", "Donator", "Tokens", "USD"],
            lambda e: [[e["username"], r.get("date",""), r.get("don_name",""),
                        r.get("tokens",""), f"${r.get('usd','')}"]
                       for r in (e["data"].get("tables",{}).get("top100tips") or [])])

    return wb


def _get_tip_rows(details):
    if not details:
        return []
    if isinstance(details, dict):
        return details.get("rows", []) or []
    return details or []


def _write_transposed_section(ws, title, headers, data_rows, start_row):
    """Write a section transposed: col A = field label, cols B+ = values per entry."""
    if not data_rows:
        return start_row

    # Section title
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    start_row += 1

    for hi, header in enumerate(headers):
        row_idx = start_row + hi
        # Label cell
        label_cell = ws.cell(row=row_idx, column=1, value=header)
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = PatternFill("solid", fgColor="D0D7E3")
        label_cell.border = THIN
        # Value cells
        for ci, row_data in enumerate(data_rows):
            val = row_data[hi] if hi < len(row_data) else ""
            cell = ws.cell(row=row_idx, column=2 + ci, value=val)
            cell.fill = ROW_ODD_FILL if ci % 2 == 0 else ROW_EVEN_FILL
            cell.font = BODY_FONT
            cell.border = THIN

    return start_row + len(headers) + 1  # +1 blank gap


def _build_per_model_workbook(all_data, platform, sections=None):
    """One sheet per model, all sections transposed (fields as rows, entries as columns)."""
    ALL = {"profile","last30d","recent_tips","top_monthly","top_alltime","biggest_tips"}
    inc = sections if sections else ALL
    platform_name = PLATFORM_NAMES.get(str(platform), f"Platform {platform}")
    wb = Workbook()
    wb.remove(wb.active)

    for entry in all_data:
        username = entry["username"]
        data = entry["data"]
        p = data.get("profile", {})
        tables = data.get("tables", {})

        safe_name = username[:31].replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "")
        ws = wb.create_sheet(title=safe_name)
        ws.column_dimensions["A"].width = 22
        cur_row = 1

        if "profile" in inc:
            cur_row = _write_transposed_section(ws, "PROFILE",
                ["Username", "Platform", "Status", "Type", "Last Online", "Last Month USD", "All Time USD"],
                [[username, platform_name, data.get("status",""), p.get("type",""),
                  p.get("last_online",""), p.get("last_month_usd",""), p.get("all_time_usd","")]],
                cur_row)

        if "last30d" in inc:
            income = tables.get("searchIncome") or []
            if income:
                cur_row = _write_transposed_section(ws, "LAST 30 DAYS",
                    ["Date", "Dons", "Tips", "Avg USD", "Total USD"],
                    [[r.get("date",""), r.get("dons",""), r.get("tips",""),
                      r.get("avg_usd",""), r.get("total_usd","")] for r in income],
                    cur_row)

        if "recent_tips" in inc:
            tips = _get_tip_rows(tables.get("incomeDetails"))
            if tips:
                cur_row = _write_transposed_section(ws, "RECENT TIPS",
                    ["Date", "Donator", "Tokens", "USD"],
                    [[r.get("date",""), r.get("don_name",""), r.get("tokens",""), r.get("usd","")] for r in tips],
                    cur_row)

        if "top_monthly" in inc:
            mons = tables.get("donsMonth") or []
            if mons:
                cur_row = _write_transposed_section(ws, "TOP MONTHLY TIPPERS",
                    ["Rank", "Donator", "Tokens", "USD"],
                    [[i+1, r.get("name",""),
                      round(r.get("total_usd",0)/TOKEN_COST) if r.get("total_usd") else "",
                      r.get("total_usd","")] for i, r in enumerate(mons)],
                    cur_row)

        if "top_alltime" in inc:
            alltime = tables.get("donsAll") or []
            if alltime:
                cur_row = _write_transposed_section(ws, "TOP ALL-TIME TIPPERS",
                    ["Rank", "Donator", "Tokens", "USD"],
                    [[i+1, r.get("name",""),
                      round(r.get("total_usd",0)/TOKEN_COST) if r.get("total_usd") else "",
                      r.get("total_usd","")] for i, r in enumerate(alltime)],
                    cur_row)

        if "biggest_tips" in inc:
            bigtips = tables.get("top100tips") or []
            if bigtips:
                cur_row = _write_transposed_section(ws, "BIGGEST TIPS",
                    ["Date", "Donator", "Tokens", "USD"],
                    [[r.get("date",""), r.get("don_name",""), r.get("tokens",""), r.get("usd","")] for r in bigtips],
                    cur_row)

        auto_width(ws)

    return wb


def build_excel_to_file(all_data, platform, path, sections=None):
    wb = _build_excel_workbook(all_data, platform, sections)
    wb.save(path)

def build_excel(all_data, platform, sections=None):
    buf = io.BytesIO()
    _build_excel_workbook(all_data, platform, sections).save(buf)
    buf.seek(0)
    return buf

def build_per_model_excel_to_file(all_data, platform, path, sections=None):
    wb = _build_per_model_workbook(all_data, platform, sections)
    wb.save(path)

def build_per_model_excel(all_data, platform, sections=None):
    buf = io.BytesIO()
    _build_per_model_workbook(all_data, platform, sections).save(buf)
    buf.seek(0)
    return buf

@app.route("/api/health")
def health():
    return jsonify({"ok": True})


@app.route("/api/scrape", methods=["POST"])
def start_scrape():
    body = request.json or {}
    platform = str(body.get("platform", "3"))
    usernames_raw = body.get("usernames", "")
    delay = float(body.get("delay", 3.0))
    name = body.get("name", "").strip()

    usernames = [
        line.strip()
        for line in usernames_raw.splitlines()
        if line.strip() and not line.startswith("#")
    ]

    if not usernames:
        return jsonify({"error": "No usernames provided"}), 400
    if platform not in PLATFORM_NAMES:
        return jsonify({"error": "Invalid platform"}), 400

    job_id = str(uuid.uuid4())
    job = {
        "id": job_id,
        "name": name,
        "platform": platform,
        "status": "queued",
        "total": len(usernames),
        "done": 0,
        "results": [],
        "log": [],
        "all_usernames": usernames,
        "created_at": datetime.datetime.utcnow().isoformat(),
    }
    save_job(job)

    thread = threading.Thread(
        target=run_scrape_job,
        args=(job_id, platform, usernames, delay),
        daemon=True,
    )
    thread.start()

    return jsonify({"job_id": job_id})


@app.route("/api/jobs/<job_id>/abandon", methods=["POST"])
def abandon_job(job_id):
    """Force-mark a stuck/orphaned job as stopped so it can be downloaded and resumed."""
    job = load_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    job["status"] = "stopped"
    job["log"].append("  ⚠ marked stopped (orphaned after redeploy)")
    save_job(job)
    return jsonify({"ok": True, "done": job["done"]})


@app.route("/api/jobs/<job_id>/stop", methods=["POST"])
def stop_job(job_id):
    job = load_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    if job["status"] not in ("running", "queued"):
        return jsonify({"error": "Job is not running"}), 400
    flag = _stop_flags.get(job_id)
    if flag:
        flag.set()
    return jsonify({"ok": True})


@app.route("/api/jobs/<job_id>/resume", methods=["POST"])
def resume_job(job_id):
    old_job = load_job(job_id)
    if not old_job:
        return jsonify({"error": "Job not found"}), 404
    if old_job["status"] not in ("error", "stopped"):
        return jsonify({"error": "Only errored or stopped jobs can be resumed"}), 400

    body = request.json or {}
    delay = float(body.get("delay", 3.0))

    completed = completed_usernames(job_id)
    all_usernames = old_job.get("all_usernames", [])
    remaining = [u for u in all_usernames if u not in completed]

    if not remaining:
        return jsonify({"error": "No remaining usernames to scrape"}), 400

    new_job_id = str(uuid.uuid4())
    new_job = {
        "id": new_job_id,
        "name": old_job.get("name", ""),
        "platform": old_job["platform"],
        "status": "queued",
        "total": len(all_usernames),
        "done": len(completed),
        "log": [f"  ↩ resumed from job {job_id[:8]}... ({len(completed)} done, {len(remaining)} remaining)"],
        "all_usernames": all_usernames,
        "resumed_from": job_id,
        "created_at": datetime.datetime.utcnow().isoformat(),
    }
    save_job(new_job)

    # Symlink or copy old results dir so new job can access them
    old_results = JOBS_DIR / f"{job_id}_results"
    new_results = JOBS_DIR / f"{new_job_id}_results"
    if old_results.exists():
        import shutil
        shutil.copytree(str(old_results), str(new_results))

    thread = threading.Thread(
        target=run_scrape_job,
        args=(new_job_id, old_job["platform"], remaining, delay),
        daemon=True,
    )
    thread.start()

    return jsonify({"job_id": new_job_id})


@app.route("/api/jobs/<job_id>")
def job_status(job_id):
    job = load_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    completed = completed_usernames(job_id)
    all_usernames = job.get("all_usernames", [])
    remaining = [u for u in all_usernames if u not in completed]

    return jsonify({
        "id": job["id"],
        "name": job.get("name", ""),
        "status": job["status"],
        "total": job["total"],
        "done": job["done"],
        "log": job["log"][-50:],
        "error": job.get("error"),
        "resumed_from": job.get("resumed_from"),
        "remaining": remaining if job["status"] in ("error", "stopped") else [],
    })


@app.route("/api/jobs/<job_id>/download/<fmt>")
def download(job_id, fmt):
    job = load_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    if job["status"] not in ("done", "stopped"):
        return jsonify({"error": "Job not complete"}), 400

    all_data = load_results(job_id)
    if not all_data:
        return jsonify({"error": "No results found"}), 404

    platform = job["platform"]
    platform_name = PLATFORM_NAMES.get(platform, "data")
    ts = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M")
    custom_name = job.get("name", "")
    base_name = custom_name if custom_name else f"statbate_{platform_name}_{ts}"
    base_name = re.sub(r'[^\w\-. ]', '_', base_name).strip()

    # Parse requested sections
    sections_param = request.args.get("sections", "")
    sections = set(sections_param.split(",")) if sections_param else None

    # Write to a temp file on disk instead of BytesIO to avoid RAM spike
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=JOBS_DIR)
    tmp.close()

    try:
        if fmt == "xlsx":
            build_excel_to_file(all_data, platform, tmp.name, sections)
            filename = f"{base_name}.xlsx"
        elif fmt == "csv":
            build_per_model_excel_to_file(all_data, platform, tmp.name, sections)
            filename = f"{base_name}_per_model.xlsx"
        else:
            return jsonify({"error": "Invalid format"}), 400

        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return send_file(tmp.name, mimetype=mimetype,
                         as_attachment=True, download_name=filename)
    except Exception as e:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass
        raise e


@app.route("/api/jobs")
def list_jobs():
    return jsonify([
        {
            "id": j["id"],
            "name": j.get("name", ""),
            "status": j["status"],
            "platform": PLATFORM_NAMES.get(j["platform"], j["platform"]),
            "total": j["total"],
            "done": j["done"],
            "created_at": j.get("created_at"),
            "resumed_from": j.get("resumed_from"),
        }
        for j in list_jobs_from_disk()
    ])


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
